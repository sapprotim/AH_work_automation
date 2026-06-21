# Add the User ID to the userId Excel file.
# Update the table row and table data as per the calendar date required data.

import os
import time
from time import sleep
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

global email, password, otp
email = os.environ.get("GH_EMAIL", "")
password = os.environ.get("GH_PASSWORD", "")
otp = os.environ.get("GH_OTP", "")

# Setup Selenium WebDriver
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://gethealthier.connectedlife.io/")
driver.implicitly_wait(100)

try:
    def start_login_and_submit_otp(driver, email, password, otp):
        driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
        driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
        driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
        driver.find_element(By.XPATH, "//input[@value='Submit']").click()
        time.sleep(5)

    def logout_login(driver, email, password, otp):
        try:
            driver.find_element(By.XPATH, "//img[@class='nameImage']").click()
            driver.find_element(By.XPATH, "//a[normalize-space()='Logout']").click()
            driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
            time.sleep(3)
            driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
            driver.find_element(By.XPATH, "//input[@value='Submit']").click()
            time.sleep(5)
        except NoSuchElementException:
            login_and_submit_otp(driver, email, password, otp)


    def login_and_submit_otp(driver, email, password, otp):
        try:
            driver.find_element(By.XPATH, "//input[@name='username']").send_keys(email)
            driver.find_element(By.XPATH, "//input[@name='password']").send_keys(password)
            driver.find_element(By.XPATH, "//input[@value='Sign In']").click()
            time.sleep(3)
            driver.find_element(By.XPATH, "//input[@id='inp']").send_keys(otp)
            driver.find_element(By.XPATH, "//input[@value='Submit']").click()
            time.sleep(5)
        except NoSuchElementException:
            driver.quit()
            driver = webdriver.Chrome()
            driver.maximize_window()
            driver.get("https://gethealthier.connectedlife.io/")
            driver.implicitly_wait(100)
            login_and_submit_otp(driver, email, password, otp)

    def Search_User(driver, username):
        try:
            time.sleep(3)
            try:
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").send_keys(username)
                time.sleep(2)
                driver.find_element(By.XPATH, "//div[@class='blurry-text ng-star-inserted']").click()
            except NoSuchElementException:
                driver.refresh()
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").clear()
                driver.find_element(By.XPATH, "//input[@placeholder='Search']").send_keys(username)
                time.sleep(2)
                driver.find_element(By.XPATH, "//div[@class='blurry-text ng-star-inserted']").click()
            time.sleep(2)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)

    def select_date_overview(driver):
        try:
            driver.find_element(By.XPATH, "//input[@id='overview-datepicker']").click()
            # When need seleted previous month (1st line only)
            driver.find_element(By.XPATH, "//th[@class='prev available']//span").click()
            driver.find_element(By.XPATH, "//div[@class='drp-calendar left single']//div[@class='calendar-table']//tr[5]/td[4]").click()
            time.sleep(2)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)

    def get_collect_data(driver, sheet, row):
        try:
            # Height (cms)
            height = driver.find_element(By.XPATH, "//p[@id='user-height']").text
            height = height.replace(" cm", "")
            if height == "--":
                height = "Null"
            sheet.cell(row=row, column=16).value = height
            print("Height", height)

            # weight (kg)
            weight_element = driver.find_element(By.XPATH, '//*[@id="profile-weight"]/p[2]')
            weight = weight_element.text
            weight = weight.replace(" Kg", "")

            if weight == "--":
                weight = "Null"

            sheet.cell(row=row, column=15).value = weight
            print("weight", weight)


        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)



    def print_user(username):
        try:
            driver.find_element(By.XPATH, "//span[normalize-space()='My Users']").click()
            print(username)
        except NoSuchElementException:
            logout_login(driver, email, password, otp)
            Search_User(driver, username)
            select_date_overview(driver)
            get_collect_data(driver, sheet, row)
            print_user(username)

    file_path = 'user_id.xlsx'
    df = pd.read_excel(file_path)
    usernames = df['User Name'].unique()
    usernames_list = usernames.tolist()

    # Load existing workbook
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    row = 2

    start_login_and_submit_otp(driver, email, password, otp)
    for username in usernames_list:
        Search_User(driver, username)
        select_date_overview(driver)
        get_collect_data(driver, sheet, row)
        print_user(username)
        row += 1

    workbook.save(file_path)
    print(f"File saved successfully at {file_path}")
except:
    workbook.save(file_path)
    print(f" Have same error but File saved successfully at {file_path}")